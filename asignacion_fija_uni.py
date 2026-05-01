import re
import streamlit as st
import pandas as pd
import requests as _requests
import unicodedata
import io
import json
import os
from datetime import date, datetime
from utils import (
    render_header, render_guide, render_stat, render_label,
    render_tip, render_error_item,
    create_progress_tracker, update_progress, finish_progress,
)

_LOADER_CSS = """
<style>
@keyframes sr-afu-spin { to { transform: rotate(360deg); } }
@keyframes sr-afu-pulse { 0%,100% { opacity:1; } 50% { opacity:.5; } }
.sr-afu-loader {
    display: flex; flex-direction: column; align-items: center;
    padding: 2rem 1rem; margin: 1rem 0; border-radius: 0.8rem;
    background: linear-gradient(135deg, rgba(42,43,161,0.08) 0%, rgba(54,156,255,0.08) 100%);
    border: 1px solid rgba(42,43,161,0.15);
}
.sr-afu-spinner {
    width: 56px; height: 56px; border-radius: 50%;
    border: 5px solid rgba(42,43,161,0.15);
    border-top-color: #2A2BA1; border-right-color: #369CFF;
    animation: sr-afu-spin 0.9s linear infinite;
}
.sr-afu-text {
    margin-top: 1rem; color: #2A2BA1; font-weight: 700;
    font-size: 1rem; letter-spacing: -0.01em;
    animation: sr-afu-pulse 1.4s ease-in-out infinite;
}
.sr-afu-sub { margin-top: 0.25rem; color: #666; font-size: 0.8rem; }
</style>
"""


def _render_loader(placeholder, mensaje, sub=""):
    sub_html = f'<div class="sr-afu-sub">{sub}</div>' if sub else ""
    placeholder.markdown(
        f'{_LOADER_CSS}<div class="sr-afu-loader">'
        f'<div class="sr-afu-spinner"></div>'
        f'<div class="sr-afu-text">{mensaje}</div>{sub_html}</div>',
        unsafe_allow_html=True,
    )

AGENCIAS_VALIDAS = {"tlahuac": "Tláhuac", "monterrey": "Monterrey"}
HORA_INICIO_FIJA = "07:00"
HORA_FINAL_FIJA = "23:00"
DURACION_FIJA = 7
UPSERT_BATCH_SIZE = 500


def _col_letter_to_index(letter):
    letter = letter.upper()
    result = 0
    for c in letter:
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result - 1


COL_AGENCIA = _col_letter_to_index("C")
COL_CLIENTE = _col_letter_to_index("D")
COL_SECTOR = _col_letter_to_index("AH")


def _sin_acentos(texto):
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _normalizar_agencia(valor):
    clave = _sin_acentos(str(valor).strip()).lower()
    return AGENCIAS_VALIDAS.get(clave)


def _get_supabase_client():
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
    except (KeyError, FileNotFoundError):
        st.error("Faltan credenciales de Supabase en secrets. Agregar [supabase] con url y key.")
        st.stop()
    from supabase import create_client
    return create_client(url, key)


def _leer_archivo(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        df = pd.read_csv(archivo, dtype=str, header=0, encoding="ISO-8859-1")
    else:
        df = pd.read_excel(archivo, dtype=str, header=0)
    return df.fillna("")


def _extraer_registros(df):
    registros = []
    descartados_agencia = 0
    descartados_vacios = 0

    max_col = max(COL_AGENCIA, COL_CLIENTE, COL_SECTOR)
    for _, row in df.iterrows():
        if len(row) <= max_col:
            descartados_vacios += 1
            continue

        agencia_raw = str(row.iloc[COL_AGENCIA]).strip()
        cliente = str(row.iloc[COL_CLIENTE]).strip()
        sector = str(row.iloc[COL_SECTOR]).strip()

        if not cliente:
            descartados_vacios += 1
            continue

        agencia = _normalizar_agencia(agencia_raw)
        if agencia is None:
            descartados_agencia += 1
            continue

        registros.append({
            "cliente": cliente,
            "sector": sector,
            "agencia": agencia,
        })

    vistos = set()
    unicos = []
    duplicados = 0
    for r in registros:
        if r["cliente"] in vistos:
            duplicados += 1
            continue
        vistos.add(r["cliente"])
        unicos.append(r)

    return unicos, {
        "descartados_agencia": descartados_agencia,
        "descartados_vacios": descartados_vacios,
        "duplicados": duplicados,
    }


def _contar_existentes(supabase, clientes):
    existentes = set()
    for i in range(0, len(clientes), 500):
        lote = clientes[i:i + 500]
        try:
            resp = supabase.table("planeacion_nacional").select("cliente").in_("cliente", lote).execute()
            for row in resp.data or []:
                existentes.add(row["cliente"])
        except Exception as e:
            st.warning(f"No se pudo contar existentes: {e}")
            return None
    return existentes


def _upsert_lote(supabase, registros):
    return supabase.table("planeacion_nacional").upsert(
        registros, on_conflict="cliente"
    ).execute()


_META_FILE = os.path.join(os.path.dirname(__file__), ".planeacion_meta.json")


def _get_last_updated() -> dict | None:
    try:
        with open(_META_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def _set_last_updated(exitosos: int, total: int):
    data = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "exitosos": exitosos,
        "total": total,
    }
    try:
        with open(_META_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass


def _rotar_habilidades(existentes, nueva):
    """
    Pone `nueva` en posicion 1 y recorre las demas sin duplicados.
    Trata el mismo numero con o sin prefijo F como el mismo skill
    (ej: '20020' y 'F20020' se consideran iguales) para migrar el
    formato viejo al nuevo de forma transparente.
    existentes: lista de 4 valores (pueden ser None/vacios).
    Retorna: lista de exactamente 4 valores.
    """
    nueva_num = nueva.lstrip("F")
    limpia = [
        h for h in existentes
        if h and str(h).strip() not in ("", "None", "null", "nan")
    ]
    limpia = [h for h in limpia if str(h).lstrip("F") != nueva_num]
    rotada = ([nueva] + limpia)[:4]
    while len(rotada) < 4:
        rotada.append(None)
    return rotada


def _seccion_actualizar_planeacion():
    meta = _get_last_updated()
    if meta:
        render_tip(
            f"Última actualización: <strong>{meta['ts']}</strong> — "
            f"{meta['exitosos']:,} de {meta['total']:,} registros subidos"
        )
    else:
        render_tip("Sin actualizaciones registradas aún.")

    render_label("Archivo Excel de planeacion")
    archivo = st.file_uploader(
        "Sube el archivo con la planeacion",
        type=["xlsx", "xls", "csv"],
        key="afu_archivo",
    )

    if not archivo:
        render_tip(
            "Columnas esperadas: <strong>C</strong> = Agencia, <strong>D</strong> = Cliente, "
            "<strong>AH</strong> = Sector. "
            "Solo se cargan filas de <strong>Tláhuac</strong> y <strong>Monterrey</strong>."
        )
        return

    loader = st.empty()
    _render_loader(loader, "Leyendo archivo...", f"Procesando {archivo.name}")
    try:
        df = _leer_archivo(archivo)
    except Exception as e:
        loader.empty()
        st.error(f"Error al leer el archivo: {e}")
        return

    _render_loader(loader, "Extrayendo registros validos...", f"{len(df):,} filas detectadas")
    registros, stats = _extraer_registros(df)
    loader.empty()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(render_stat(len(df), "Filas leidas"), unsafe_allow_html=True)
    with col2:
        n_tlahuac = sum(1 for r in registros if r["agencia"] == "Tláhuac")
        st.markdown(render_stat(n_tlahuac, "Tláhuac"), unsafe_allow_html=True)
    with col3:
        n_mty = sum(1 for r in registros if r["agencia"] == "Monterrey")
        st.markdown(render_stat(n_mty, "Monterrey"), unsafe_allow_html=True)
    with col4:
        st.markdown(render_stat(len(registros), "A subir"), unsafe_allow_html=True)

    if stats["descartados_agencia"] or stats["descartados_vacios"] or stats["duplicados"]:
        render_tip(
            f"Descartes — otras agencias: <strong>{stats['descartados_agencia']}</strong> · "
            f"vacios: <strong>{stats['descartados_vacios']}</strong> · "
            f"duplicados: <strong>{stats['duplicados']}</strong>"
        )

    if not registros:
        render_tip("No hay filas validas para subir.", warning=True)
        return

    render_label("Vista previa")
    st.dataframe(pd.DataFrame(registros).head(20), use_container_width=True)

    if not st.button("Subir a Supabase", type="primary", use_container_width=True):
        return

    supabase = _get_supabase_client()
    loader = st.empty()

    _render_loader(loader, "Consultando clientes existentes...", f"{len(registros):,} a verificar")
    clientes = [r["cliente"] for r in registros]
    existentes = _contar_existentes(supabase, clientes)

    total = len(registros)
    total_lotes = (total + UPSERT_BATCH_SIZE - 1) // UPSERT_BATCH_SIZE
    barra, contador, contenedor_errores = create_progress_tracker(total, text="Subiendo a Supabase...")

    procesados = 0
    errores = 0
    for i in range(0, total, UPSERT_BATCH_SIZE):
        lote_num = i // UPSERT_BATCH_SIZE + 1
        _render_loader(loader, f"Subiendo lote {lote_num} de {total_lotes}...", f"{len(registros[i:i+UPSERT_BATCH_SIZE])} registros en este lote")
        lote = registros[i:i + UPSERT_BATCH_SIZE]
        try:
            _upsert_lote(supabase, lote)
        except Exception as e:
            errores += len(lote)
            with contenedor_errores:
                render_error_item(f"Lote {lote_num}: {e}")
        procesados += len(lote)
        update_progress(barra, contador, procesados, total, text="Subiendo a Supabase...")

    loader.empty()
    finish_progress(barra)

    exitosos = total - errores
    nuevos = exitosos - len(existentes) if existentes is not None else None
    actualizados = len(existentes) if existentes is not None else None

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(render_stat(exitosos, "Procesados OK"), unsafe_allow_html=True)
    with col2:
        if nuevos is not None:
            st.markdown(render_stat(nuevos, "Nuevos"), unsafe_allow_html=True)
    with col3:
        if actualizados is not None:
            st.markdown(render_stat(actualizados, "Actualizados"), unsafe_allow_html=True)

    if errores:
        render_tip(f"Hubo {errores} registros con error. Revisa los mensajes arriba.", warning=True)
    else:
        render_tip("Todos los registros se subieron correctamente.")
        _set_last_updated(exitosos, total)


HAB_COL_B_HABILIDAD = 1
HAB_COL_S_CLIENTE = 18

RUTEO_COL_D_HORA_INICIAL = 3
RUTEO_COL_E_HORA_FINAL = 4
RUTEO_COL_F_TIEMPO_SERVICIO = 5
RUTEO_COL_G_NOTA = 6
RUTEO_COL_J_REFERENCE = 9
RUTEO_COL_K_HABILIDADES = 10
RUTEO_COL_Q = 16
RUTEO_COL_R = 17



def _fetch_planeacion(supabase, clientes):
    datos = {}
    for i in range(0, len(clientes), 500):
        lote = clientes[i:i + 500]
        try:
            resp = supabase.table("planeacion_nacional").select(
                "cliente,hora_inicio,hora_final,duracion,"
                "habilidad_1,habilidad_2,habilidad_3,habilidad_4"
            ).in_("cliente", lote).execute()
            for row in resp.data or []:
                datos[row["cliente"]] = row
        except Exception as e:
            st.warning(f"Error consultando Supabase: {e}")
            return datos
    return datos


def _procesar_ruteo(df, nombre_original, habilidades_disponibles, agencia="Tláhuac"):
    loader = st.empty()
    supabase = _get_supabase_client()

    _render_loader(loader, "Consultando Supabase...", "Buscando coincidencias por nota")
    notas = [str(v).strip() for v in df.iloc[:, RUTEO_COL_G_NOTA]]
    notas_unicas = list({n for n in notas if n})
    lookup = _fetch_planeacion(supabase, notas_unicas)

    total = len(df)
    _render_loader(loader, "Actualizando filas...", f"{total:,} filas, {len(lookup):,} clientes con match")

    matched = 0
    unmatched = 0
    ruteo_dia_filas = []
    for idx in range(total):
        nota = str(df.iat[idx, RUTEO_COL_G_NOTA]).strip()
        data = lookup.get(nota) if nota else None

        # Leer valores originales del archivo antes de sobreescribir
        hora_i_orig = str(df.iat[idx, RUTEO_COL_D_HORA_INICIAL]).strip() if df.shape[1] > RUTEO_COL_D_HORA_INICIAL else ""
        hora_f_orig = str(df.iat[idx, RUTEO_COL_E_HORA_FINAL]).strip() if df.shape[1] > RUTEO_COL_E_HORA_FINAL else ""
        dur_orig = str(df.iat[idx, RUTEO_COL_F_TIEMPO_SERVICIO]).strip() if df.shape[1] > RUTEO_COL_F_TIEMPO_SERVICIO else ""

        if data:
            matched += 1
            hora_i = data.get("hora_inicio") or "07:00"
            hora_f = data.get("hora_final") or "23:00"
            dur = data.get("duracion") if data.get("duracion") is not None else 7
            df.iat[idx, RUTEO_COL_D_HORA_INICIAL] = hora_i
            df.iat[idx, RUTEO_COL_E_HORA_FINAL] = hora_f
            df.iat[idx, RUTEO_COL_F_TIEMPO_SERVICIO] = dur
            hab_asignada = "Fuera"
            for n in range(1, 5):
                h = (data.get(f"habilidad_{n}") or "").strip()
                if h and h in habilidades_disponibles:
                    hab_asignada = h
                    break
            df.iat[idx, RUTEO_COL_K_HABILIDADES] = hab_asignada
        else:
            unmatched += 1
            df.iat[idx, RUTEO_COL_D_HORA_INICIAL] = "07:00"
            df.iat[idx, RUTEO_COL_E_HORA_FINAL] = "23:00"
            df.iat[idx, RUTEO_COL_F_TIEMPO_SERVICIO] = 7
            df.iat[idx, RUTEO_COL_K_HABILIDADES] = "Fuera"

        # Recolectar para ruteo_dia con los valores originales del archivo
        ref = str(df.iat[idx, RUTEO_COL_J_REFERENCE]).strip() if df.shape[1] > RUTEO_COL_J_REFERENCE else ""
        if ref and ref not in ("", "nan", "None"):
            carga_2 = str(df.iat[idx, RUTEO_COL_Q]).strip() if df.shape[1] > RUTEO_COL_Q else ""
            carga_3 = str(df.iat[idx, RUTEO_COL_R]).strip() if df.shape[1] > RUTEO_COL_R else ""
            ruteo_dia_filas.append({
                "reference": ref,
                "hora_inicio": hora_i_orig if hora_i_orig not in ("", "nan", "None") else "07:00",
                "hora_final": hora_f_orig if hora_f_orig not in ("", "nan", "None") else "23:00",
                "duracion": dur_orig if dur_orig not in ("", "nan", "None") else "7",
                "carga_2": carga_2 if carga_2 not in ("", "nan", "None") else None,
                "carga_3": carga_3 if carga_3 not in ("", "nan", "None") else None,
            })

        df.iat[idx, RUTEO_COL_Q] = ""
        df.iat[idx, RUTEO_COL_R] = ""

    if ruteo_dia_filas:
        _guardar_ruteo_dia(supabase, ruteo_dia_filas, agencia)

    _render_loader(loader, "Generando archivo...", "Escribiendo xlsx")

    for col_idx in [2, 5, 6, 7, 8, 13]:
        if col_idx < df.shape[1]:
            df.iloc[:, col_idx] = pd.to_numeric(df.iloc[:, col_idx], errors="coerce")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ruteo")
        from openpyxl.styles import Border, Font, PatternFill

        ws = writer.sheets["Ruteo"]

        no_border = Border()
        sin_fondo = PatternFill(fill_type=None)
        for cell in ws[1]:
            cell.border = no_border
            cell.fill = sin_fondo
            cell.font = Font(bold=False)

        formatos = {
            3: "0",        # C
            6: "0",        # F
            7: "0",        # G
            8: "0.00000",  # H
            9: "0.00000",  # I
            14: "0",       # N
        }
        for col_1idx, fmt in formatos.items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_1idx).number_format = fmt

    buffer.seek(0)
    loader.empty()

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(render_stat(total, "Filas procesadas"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_stat(matched, "Con match"), unsafe_allow_html=True)
    with col3:
        st.markdown(render_stat(unmatched, 'Sin match ("Fuera")'), unsafe_allow_html=True)

    nombre_out = nombre_original.rsplit(".", 1)[0] + "_actualizado.xlsx"
    st.download_button(
        "Descargar archivo actualizado",
        buffer,
        file_name=nombre_out,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


def _tabla_ruteo_dia(agencia):
    return "ruteo_dia_tlahuac" if agencia == "Tláhuac" else "ruteo_dia_monterrey"


def _guardar_ruteo_dia(supabase, filas, agencia):
    tabla = _tabla_ruteo_dia(agencia)
    try:
        supabase.table(tabla).upsert(filas, on_conflict="reference").execute()
    except Exception as e:
        st.warning(f"No se pudo guardar en {tabla}: {e}")


def _enviar_actualizaciones(token, visitas_put):
    from concurrent.futures import ThreadPoolExecutor, as_completed

    MAX_BLOCK = 400
    bloques = [visitas_put[i:i + MAX_BLOCK] for i in range(0, len(visitas_put), MAX_BLOCK)]
    total_bloques = len(bloques)
    headers = {"Authorization": f"Token {token}", "Content-Type": "application/json"}

    def _put_lote(lote):
        r = _requests.put(
            "https://api.simpliroute.com/v1/routes/visits/",
            json=lote, headers=headers, timeout=120,
        )
        return len(lote), r.status_code, r.text[:200]

    barra = st.progress(0, text=f"Enviando {len(visitas_put)} visitas en {total_bloques} bloques...")
    resultados = {}

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(_put_lote, lote): i for i, lote in enumerate(bloques)}
        completados = 0
        for future in as_completed(futures):
            i = futures[future]
            try:
                n, status, text = future.result()
                resultados[i] = (n, status, text)
            except Exception as e:
                resultados[i] = (len(bloques[i]), 0, str(e))
            completados += 1
            barra.progress(completados / total_bloques, text=f"Bloques completados: {completados}/{total_bloques}")

    barra.empty()

    ok = sum(n for n, status, _ in resultados.values() if status in (200, 201, 204))
    errores_put = [
        f"Bloque {i + 1}: HTTP {status} — {text}"
        for i, (n, status, text) in sorted(resultados.items())
        if status not in (200, 201, 204)
    ]

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(render_stat(ok, "Actualizados OK"), unsafe_allow_html=True)
    with col2:
        if errores_put:
            st.markdown(
                render_stat(len(errores_put), "Bloques con error",
                            style="background: linear-gradient(135deg, #d32f2f 0%, #b71c1c 100%);"),
                unsafe_allow_html=True,
            )
    for e in errores_put:
        render_error_item(e)
    if not errores_put:
        render_tip("Todos los registros actualizados correctamente.")


def _try_num(val):
    if val is None:
        return None
    try:
        return float(val) if "." in str(val) else int(val)
    except (ValueError, TypeError):
        return None


def _seccion_actualizar_datos_simpli():
    render_label("Cuenta")
    cuenta = st.radio(
        "Cuenta",
        ["Tláhuac", "Monterrey"],
        horizontal=True,
        key="ads_cuenta",
        label_visibility="collapsed",
    )
    token_key = "token_tlahuac" if cuenta == "Tláhuac" else "token_monterrey"
    try:
        token = st.secrets["cuentas_unilever"][token_key].strip()
    except Exception:
        render_tip("Token no configurado en secrets.", warning=True)
        return

    render_label("Fecha de entrega")
    fecha = st.date_input(
        "Fecha",
        value=date.today(),
        key="ads_fecha",
        format="DD/MM/YYYY",
        label_visibility="collapsed",
    )
    fecha_str = fecha.strftime("%Y-%m-%d")

    if st.button("Consultar visitas", use_container_width=True, key="ads_btn_consultar"):
        st.session_state.pop("ads_visitas_put", None)
        st.session_state.pop("ads_total_visitas", None)

        loader = st.empty()
        headers = {"Authorization": f"Token {token}"}

        _render_loader(loader, "Consultando visitas en SimpliRoute...", fecha_str)
        try:
            r = _requests.get(
                f"https://api.simpliroute.com/v1/routes/visits/?planned_date={fecha_str}",
                headers=headers, timeout=60,
            )
            r.raise_for_status()
            visitas = r.json()
        except Exception as e:
            loader.empty()
            st.error(f"Error al consultar SimpliRoute: {e}")
            return
        loader.empty()

        if not isinstance(visitas, list):
            st.error(f"Respuesta inesperada de la API: {str(visitas)[:200]}")
            return

        referencias_visitas = [v.get("reference", "") for v in visitas if v.get("reference")]

        supabase = _get_supabase_client()
        lookup_ruteo = {}
        for i in range(0, len(referencias_visitas), 500):
            lote = referencias_visitas[i:i + 500]
            try:
                resp = supabase.table(_tabla_ruteo_dia(cuenta)).select(
                    "reference,hora_inicio,hora_final,duracion,carga_2,carga_3"
                ).in_("reference", lote).execute()
                for row in resp.data or []:
                    lookup_ruteo[row["reference"]] = row
            except Exception as e:
                st.warning(f"Error consultando tabla de ruteo: {e}")

        visitas_put = []
        for v in visitas:
            ref = v.get("reference", "")
            if not ref or ref not in lookup_ruteo:
                continue
            ruteo = lookup_ruteo[ref]
            visitas_put.append({
                "id": v["id"],
                "reference": ref,
                "title": v.get("title", ""),
                "address": v.get("address", ""),
                "planned_date": v.get("planned_date", ""),
                "route": v.get("route", ""),
                "window_start": ruteo.get("hora_inicio") or "",
                "window_end": ruteo.get("hora_final") or "",
                "time_at_stop": _try_num(ruteo.get("duracion")),
                "load_2": _try_num(ruteo.get("carga_2")),
                "load_3": _try_num(ruteo.get("carga_3")),
            })

        st.session_state["ads_visitas_put"] = visitas_put
        st.session_state["ads_total_visitas"] = len(visitas)

    if "ads_visitas_put" not in st.session_state:
        render_tip(
            "Ingresa la fecha de entrega y pulsa <strong>Consultar visitas</strong>. "
            "Los valores de ventana y carga provienen del último archivo procesado en "
            "<em>Generar archivo de ruteo</em>."
        )
        return

    total_visitas = st.session_state["ads_total_visitas"]
    visitas_put = st.session_state["ads_visitas_put"]
    sin_datos = total_visitas - len(visitas_put)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(render_stat(total_visitas, "Visitas en la fecha"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_stat(len(visitas_put), "Con datos de ruteo"), unsafe_allow_html=True)
    with col3:
        st.markdown(render_stat(sin_datos, "Sin datos de ruteo"), unsafe_allow_html=True)

    if not visitas_put:
        render_tip(
            "Ninguna visita tiene datos en la tabla de ruteo. "
            "Primero genera el archivo de ruteo en la tab <em>Generar archivo de ruteo</em>.",
            warning=True,
        )
        return

    if sin_datos:
        render_tip(f"{sin_datos} visitas no tienen datos en ruteo y serán omitidas.", warning=True)

    if st.button("Actualizar en SimpliRoute", type="primary", use_container_width=True, key="ads_btn_actualizar"):
        _enviar_actualizaciones(token, visitas_put)


_VEHICLE_PATTERN = re.compile(r'^R(\d+)-MX\d+$')
_SR_TIMEOUT = 30


def _extraer_num_vehiculo(texto):
    """R20020-MX01 → '20020',  '20020' → '20020',  otros → None."""
    texto = texto.strip()
    m = _VEHICLE_PATTERN.match(texto)
    if m:
        return m.group(1)
    if re.match(r'^\d+$', texto):
        return texto
    return None


def _sr_headers(token):
    return {"Authorization": f"Token {token}", "Content-Type": "application/json"}


def _actualizar_skills_tlahuac(nums_activos):
    try:
        token = st.secrets["cuentas_unilever"]["token_tlahuac"].strip()
    except (KeyError, FileNotFoundError):
        st.error("Falta token_tlahuac en [cuentas_unilever] en secrets.")
        return

    loader = st.empty()
    _render_loader(loader, "Consultando vehículos y skills en SimpliRoute...")
    try:
        vehiculos = _requests.get(
            f"https://api.simpliroute.com/v1/routes/vehicles/",
            headers=_sr_headers(token), timeout=_SR_TIMEOUT,
        ).json()
        skills_raw = _requests.get(
            f"https://api.simpliroute.com/v1/routes/skills/",
            headers=_sr_headers(token), timeout=_SR_TIMEOUT,
        ).json()
    except Exception as e:
        loader.empty()
        st.error(f"Error al consultar SimpliRoute: {e}")
        return
    loader.empty()

    skill_map = {s["skill"]: s["id"] for s in skills_raw}
    fuera_id = skill_map.get("Fuera")
    if not fuera_id:
        st.error("No se encontró el skill 'Fuera' en la cuenta.")
        return

    # Solo vehiculos con patron R#####-MX##
    managed = [(v["id"], v["name"]) for v in vehiculos if _VEHICLE_PATTERN.match(v["name"])]
    total = len(managed)

    barra, contador, contenedor_errores = create_progress_tracker(total, "Actualizando skills...")

    ok = 0
    errores = 0
    for i, (vid, vname) in enumerate(managed):
        m = _VEHICLE_PATTERN.match(vname)
        num = m.group(1) if m else None

        if num and num in nums_activos:
            skill_name = f"F{num}"
            skill_id = skill_map.get(skill_name)
            if not skill_id:
                with contenedor_errores:
                    render_error_item(f"{vname} — skill {skill_name} no existe, se asigna Fuera")
                skill_id = fuera_id
        else:
            skill_id = fuera_id

        try:
            r = _requests.patch(
                f"https://api.simpliroute.com/v1/routes/vehicles/{vid}/",
                headers=_sr_headers(token),
                json={"skills": [skill_id]},
                timeout=_SR_TIMEOUT,
            )
            if 200 <= r.status_code < 300:
                ok += 1
            else:
                with contenedor_errores:
                    render_error_item(f"{vname} — HTTP {r.status_code}")
                errores += 1
        except Exception as e:
            with contenedor_errores:
                render_error_item(f"{vname} — {e}")
            errores += 1

        update_progress(barra, contador, i + 1, total)

    finish_progress(barra)

    n_activos = sum(1 for _, n in managed if (m := _VEHICLE_PATTERN.match(n)) and m.group(1) in nums_activos)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(render_stat(ok, "Actualizados OK"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_stat(n_activos, "Con habilidad F"), unsafe_allow_html=True)
    with col3:
        st.markdown(render_stat(total - n_activos, "Con Fuera"), unsafe_allow_html=True)

    if errores:
        render_tip(f"Hubo {errores} vehículos con error.", warning=True)
    else:
        render_tip("Todos los vehículos actualizados correctamente.")


def _seccion_generar_ruteo():
    render_label("Agencia")
    agencia = st.radio(
        "Agencia",
        ["Tláhuac", "Monterrey"],
        horizontal=True,
        key="agr_agencia",
        label_visibility="collapsed",
    )

    habilidades_disponibles = set()

    if agencia == "Tláhuac":
        render_label("Vehículos activos este día")
        vehiculos_txt = st.text_area(
            "Vehiculos",
            placeholder="Un vehículo por línea\nEj: R20020-MX01 ó 20020",
            key="agr_vehiculos",
            label_visibility="collapsed",
            height=120,
        )
        nums_activos = {n for line in vehiculos_txt.splitlines() if (n := _extraer_num_vehiculo(line))}
        habilidades_disponibles = {f"F{n}" for n in nums_activos}

        if nums_activos:
            render_tip(
                f"<strong>{len(nums_activos)}</strong> vehículos activos — "
                "habilidades: "
                + ", ".join(f"<code>{h}</code>" for h in sorted(habilidades_disponibles))
                + " — los demás quedarán con <strong>Fuera</strong>."
            )
            if st.button("Actualizar skills en SimpliRoute", key="agr_btn_sr"):
                _actualizar_skills_tlahuac(nums_activos)
    else:
        col_a, col_b = st.columns(2)
        with col_a:
            render_label("Cuántas rutas")
            n_rutas = st.number_input(
                "Rutas",
                min_value=1,
                max_value=99,
                value=5,
                key="agr_n_rutas",
                label_visibility="collapsed",
            )
        with col_b:
            render_label("Cuentas especiales")
            especiales_txt = st.text_area(
                "Especiales",
                placeholder="Una por línea",
                key="agr_especiales",
                label_visibility="collapsed",
                height=80,
            )
        rutas = [str(i) for i in range(1, int(n_rutas) + 1)]
        especiales = [e.strip() for e in especiales_txt.splitlines() if e.strip()]
        habilidades_disponibles = set(rutas + especiales)

    st.markdown("---")

    render_label("Archivo Excel de ruteo")
    archivo = st.file_uploader(
        "Sube el archivo de ruteo a actualizar",
        type=["xlsx", "xls"],
        key="afu_ruteo_archivo",
        label_visibility="collapsed",
    )

    if not archivo:
        render_tip(
            "Se actualizan: <strong>D</strong> Hora Inicial, <strong>E</strong> Hora Final, "
            "<strong>F</strong> Tiempo Servicio, <strong>K</strong> Habilidades requeridas. "
            "Se vacian <strong>Q</strong> y <strong>R</strong>. "
            "Las columnas <strong>H</strong> y <strong>I</strong> (Latitud/Longitud) pasan intactas. "
            "En Supabase se guardan los valores <strong>originales</strong> de D/E/F/Q/R del archivo subido. "
            "Match por columna <strong>G</strong> (Nota) vs <strong>cliente</strong> de Supabase. "
            "Sin match o sin habilidad disponible: habilidad = <strong>Fuera</strong>."
        )
        return

    loader = st.empty()
    _render_loader(loader, "Leyendo archivo...", f"Procesando {archivo.name}")
    try:
        df = pd.read_excel(archivo, dtype=str, header=0).fillna("")
    except Exception as e:
        loader.empty()
        st.error(f"Error al leer el archivo: {e}")
        return
    loader.empty()

    if habilidades_disponibles:
        render_tip(
            f"<strong>{len(habilidades_disponibles)}</strong> habilidades configuradas: "
            + ", ".join(f"<code>{h}</code>" for h in sorted(habilidades_disponibles))
        )
    else:
        render_tip("No hay vehículos/rutas configurados. Todos los clientes quedarán como <strong>Fuera</strong>.", warning=True)

    st.markdown(render_stat(len(df), "Filas en archivo"), unsafe_allow_html=True)
    render_label("Vista previa (primeras 10 filas)")
    st.dataframe(df.head(10), use_container_width=True)

    if st.button("Procesar y generar archivo", type="primary", use_container_width=True):
        _procesar_ruteo(df, archivo.name, habilidades_disponibles, agencia)


def _seccion_actualizar_habilidades():
    render_label("Agencia")
    agencia = st.radio(
        "Agencia",
        ["Tláhuac", "Monterrey"],
        horizontal=True,
        key="afh_agencia",
        label_visibility="collapsed",
    )

    render_label("Fecha de entrega del ruteo")
    fecha_ruteo = st.date_input(
        "Fecha de entrega del ruteo",
        value=date.today(),
        key="afh_fecha",
        format="DD/MM/YYYY",
        label_visibility="collapsed",
    )

    render_label("Archivo SimpliRoute Plan")
    archivo = st.file_uploader(
        "Sube el archivo SimpliRoute_Plan_...",
        type=["xlsx", "xls"],
        key="afh_archivo",
        label_visibility="collapsed",
    )

    if not archivo:
        render_tip(
            "Formato esperado: <strong>SimpliRoute_Plan_YYYY-MM-DD.xlsx</strong>. "
            "Columna <strong>S</strong> = Cliente (clave de match), "
            "<strong>B</strong> = Habilidad (formato R20020-MX01 → se extrae 20020). "
            "La habilidad se coloca en <strong>habilidad_1</strong> y se recorren las existentes."
        )
        return

    loader = st.empty()
    _render_loader(loader, "Leyendo archivo...", archivo.name)
    try:
        df = pd.read_excel(archivo, dtype=str, header=0).fillna("")
    except Exception as e:
        loader.empty()
        st.error(f"Error al leer el archivo: {e}")
        return
    loader.empty()

    registros = []
    sin_cliente = 0
    for _, row in df.iterrows():
        if len(row) <= HAB_COL_S_CLIENTE:
            sin_cliente += 1
            continue
        cliente = str(row.iloc[HAB_COL_S_CLIENTE]).strip()
        if not cliente or cliente.lower() in ("nan", "none", ""):
            sin_cliente += 1
            continue
        hab_raw = str(row.iloc[HAB_COL_B_HABILIDAD]).strip() if len(row) > HAB_COL_B_HABILIDAD else ""
        habilidad = "F" + hab_raw.lstrip("R").split("-")[0] if hab_raw else ""
        registros.append({"cliente": cliente, "habilidad": habilidad})

    vistos = set()
    unicos = []
    for r in registros:
        if r["cliente"] not in vistos:
            vistos.add(r["cliente"])
            unicos.append(r)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(render_stat(len(df), "Filas en archivo"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_stat(len(unicos), "Clientes unicos"), unsafe_allow_html=True)
    with col3:
        st.markdown(
            render_stat(sin_cliente, "Sin cliente (descartados)",
                        style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);"),
            unsafe_allow_html=True,
        )

    if not unicos:
        render_tip("No se encontraron clientes validos en la columna S.", warning=True)
        return

    render_label("Vista previa")
    st.dataframe(pd.DataFrame(unicos).head(20), use_container_width=True, hide_index=True)

    if not st.button("Actualizar habilidades en Supabase", type="primary", use_container_width=True, key="afh_btn"):
        return

    supabase = _get_supabase_client()

    # Leer habilidades actuales de Supabase para aplicar rotacion
    loader2 = st.empty()
    _render_loader(loader2, "Consultando habilidades actuales...", f"{len(unicos):,} clientes")
    clientes = [r["cliente"] for r in unicos]
    existentes_map = {}
    for i in range(0, len(clientes), 500):
        lote = clientes[i:i + 500]
        try:
            resp = supabase.table("planeacion_nacional").select(
                "cliente,habilidad_1,habilidad_2,habilidad_3,habilidad_4"
            ).in_("cliente", lote).execute()
            for row in resp.data or []:
                existentes_map[row["cliente"]] = row
        except Exception as e:
            st.warning(f"No se pudieron consultar habilidades existentes: {e}")
    loader2.empty()

    # Construir payload con rotacion
    payload = []
    for r in unicos:
        cliente = r["cliente"]
        nueva_hab = r["habilidad"]
        existente = existentes_map.get(cliente, {})
        actuales = [
            existente.get("habilidad_1"),
            existente.get("habilidad_2"),
            existente.get("habilidad_3"),
            existente.get("habilidad_4"),
        ]
        rotada = _rotar_habilidades(actuales, nueva_hab)
        payload.append({
            "cliente": cliente,
            "agencia": agencia,
            "habilidad_1": rotada[0],
            "habilidad_2": rotada[1],
            "habilidad_3": rotada[2],
            "habilidad_4": rotada[3],
        })

    total = len(payload)
    total_lotes = (total + UPSERT_BATCH_SIZE - 1) // UPSERT_BATCH_SIZE
    barra, contador, contenedor_errores = create_progress_tracker(total, text="Actualizando en Supabase...")

    procesados = 0
    errores = 0
    loader3 = st.empty()
    for i in range(0, total, UPSERT_BATCH_SIZE):
        lote_num = i // UPSERT_BATCH_SIZE + 1
        lote = payload[i:i + UPSERT_BATCH_SIZE]
        _render_loader(loader3, f"Subiendo lote {lote_num} de {total_lotes}...", f"{len(lote)} registros")
        try:
            supabase.table("planeacion_nacional").upsert(lote, on_conflict="cliente").execute()
        except Exception as e:
            errores += len(lote)
            with contenedor_errores:
                render_error_item(f"Lote {lote_num}: {e}")
        procesados += len(lote)
        update_progress(barra, contador, procesados, total, text="Actualizando en Supabase...")

    loader3.empty()
    finish_progress(barra)

    exitosos = total - errores
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(render_stat(exitosos, "Actualizados OK"), unsafe_allow_html=True)
    if errores:
        with col2:
            st.markdown(
                render_stat(errores, "Con error",
                            style="background: linear-gradient(135deg, #d32f2f 0%, #b71c1c 100%);"),
                unsafe_allow_html=True,
            )
        render_tip(f"Hubo {errores} registros con error. Revisa los mensajes arriba.", warning=True)
    else:
        render_tip(
            f"Todos los registros se actualizaron correctamente — "
            f"<strong>{agencia}</strong> · {fecha_ruteo.strftime('%d/%m/%Y')}."
        )


def pagina_asignacion_fija_uni():
    render_header("Asignacion Fija Uni", "Planeacion nacional de visitas Unilever")
    render_guide(
        [
            "Selecciona la accion a ejecutar.",
            "Sube el archivo Excel con la planeacion.",
            "Se filtran solo filas de Tláhuac y Monterrey (columna C).",
            "Se extraen Cliente (D) y Sector (AH).",
            "Clientes nuevos: se crean con hora 07:00-23:00 y duracion 7 por defecto.",
            "Clientes existentes: solo se actualizan sector y agencia. El resto queda intacto.",
        ],
        "El campo <strong>habilidad</strong> se cargara desde otro archivo en un paso posterior.",
    )

    tab1, tab2, tab3, tab4 = st.tabs(["Actualizar planeacion nacional", "Generar archivo de ruteo", "Actualizar Habilidades", "Actualizar datos Simpli"])
    with tab1:
        _seccion_actualizar_planeacion()
    with tab2:
        _seccion_generar_ruteo()
    with tab3:
        _seccion_actualizar_habilidades()
    with tab4:
        _seccion_actualizar_datos_simpli()
